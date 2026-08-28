import csv
import re
from datetime import date, datetime
from typing import Iterable, Optional

from src.profitability.contracts import (
    EXPENSE,
    INCOME,
    BillingRecord,
    CashflowRecord,
    RowProblem,
)

# כותרות מקובלות בייצוא של סאמיט/ריווחית/חשבשבת/אקסל ידני.
# הכותרת מנורמלת (רווחים, גרשיים ומקפים מוסרים) לפני ההשוואה.
COLUMN_ALIASES = {
    "date": ["תאריך", "תאריךחשבונית", "תאריךהחיוב", "תאריךערך", "date", "invoicedate"],
    "client": ["לקוח", "שםלקוח", "שםהלקוח", "לכבוד", "client", "customer", "name"],
    "amount": ["סכום", "סהכ", "סהכלתשלום", "סכוםכולל", "amount", "total", "sum"],
    "amount_net": ["לפנימעמ", "סכוםלפנימעמ", "סכוםנטו", "נטו", "amountnet", "net", "subtotal"],
    "vat": ["מעמ", "סכוםמעמ", "מסערךמוסף", "vat", "tax"],
    "direction": ["סוג", "סוגתנועה", "כיוון", "חובהזכות", "direction", "type"],
    "category": ["קטגוריה", "סעיף", "סיווג", "תיאור", "category", "description"],
    "document": ["מסמך", "מספרמסמך", "חשבונית", "מסחשבונית", "document", "invoice"],
}

INCOME_WORDS = ["הכנסה", "הכנסות", "זכות", "תקבול", "תקבולים", "גבייה", "גביה", "income", "credit", "in"]
EXPENSE_WORDS = ["הוצאה", "הוצאות", "חובה", "תשלום", "תשלומים", "expense", "debit", "out"]

DATE_FORMATS = ["%d/%m/%Y", "%d.%m.%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%y", "%d.%m.%y"]


def normalize_header(header: str) -> str:
    return re.sub(r"[\s\"'`׳״\-_.]", "", (header or "").strip().lower())


def build_column_map(headers: Iterable[str]) -> dict:
    """ממפה כותרות הקובץ לשמות השדות הפנימיים. כותרת לא מוכרת פשוט מתעלמים ממנה."""
    mapping = {}
    for header in headers:
        field = match_field(normalize_header(header))
        if field and field not in mapping:
            mapping[field] = header
    return mapping


def match_field(normalized: str) -> Optional[str]:
    """התאמה מדויקת קודמת; אחרת הכינוי הארוך ביותר שמוכל בכותרת מנצח.
    כך 'סכום כולל מעמ' נקרא כסכום ברוטו, ו'סכום לפני מעמ' כנטו - ולא להפך."""
    if not normalized:
        return None
    best_field, best_length = None, 0
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias == normalized:
                return field
            if alias in normalized and len(alias) > best_length:
                best_field, best_length = field, len(alias)
    return best_field


def parse_date(value: str) -> Optional[date]:
    text = (value or "").strip()
    if not text:
        return None
    text = text.split(" ")[0].split("T")[0]
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_amount(value: str) -> Optional[float]:
    """קורא סכום בפורמט ישראלי: ₪, פסיקי אלפים, וסוגריים כמינוס."""
    text = str(value or "").strip()
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = re.sub(r"[₪$€,\s()]", "", text)
    if text in ("", "-"):
        return None
    try:
        amount = float(text)
    except ValueError:
        return None
    return -amount if negative else amount


def parse_direction(value: str) -> Optional[str]:
    text = normalize_header(value)
    if not text:
        return None
    if any(word in text for word in INCOME_WORDS):
        return INCOME
    if any(word in text for word in EXPENSE_WORDS):
        return EXPENSE
    return None


def split_vat(row: dict, columns: dict) -> Optional[tuple]:
    """מחזיר (נטו, ברוטו, האם המעמ ידוע). כלל ההכרעה מסודר מהמפורש לפחות מפורש."""
    net = parse_amount(row.get(columns["amount_net"], "")) if "amount_net" in columns else None
    gross = parse_amount(row.get(columns["amount"], "")) if "amount" in columns else None
    vat = parse_amount(row.get(columns["vat"], "")) if "vat" in columns else None

    if net is not None:
        return net, gross if gross is not None else net, True
    if gross is not None and vat is not None:
        return gross - vat, gross, True
    if gross is not None:
        # אין עמודת מעמ - הסכום נלקח כמו שהוא, והדוח יסמן שהבסיס לא מאומת.
        return gross, gross, False
    return None


def read_rows(path: str) -> tuple:
    """קורא CSV/TSV, ו-xlsx אם openpyxl מותקן. מחזיר (כותרות, שורות כמילונים)."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        return _read_xlsx(path)
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
        reader = csv.DictReader(handle, delimiter=delimiter)
        return list(reader.fieldnames or []), list(reader)


def _read_xlsx(path: str) -> tuple:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "קריאת xlsx דורשת openpyxl (pip install -r requirements-profitability.txt), "
            "או שמור את הקובץ כ-CSV"
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(cell) if cell is not None else "" for cell in next(rows, [])]
    records = []
    for values in rows:
        cells = ["" if value is None else str(value) for value in values]
        records.append(dict(zip(headers, cells)))
    workbook.close()
    return headers, records


def load_billing(path: str) -> tuple:
    """טוען קובץ חיובים. מחזיר (רשומות, שורות בעייתיות)."""
    headers, rows = read_rows(path)
    columns = build_column_map(headers)
    records, problems = [], []

    missing = [field for field in ("date", "amount") if field not in columns]
    if "amount" in missing and "amount_net" in columns:
        missing.remove("amount")
    if missing:
        raise ValueError(
            f"חסרות עמודות חובה בקובץ החיובים: {', '.join(missing)}. "
            f"הכותרות שנמצאו: {', '.join(headers)}"
        )

    for index, row in enumerate(rows, start=2):
        when = parse_date(row.get(columns["date"], ""))
        amounts = split_vat(row, columns)
        if when is None or amounts is None:
            if _is_blank(row):
                continue
            problems.append(
                RowProblem("billing", index, "תאריך או סכום לא קריאים", _preview(row))
            )
            continue
        net, gross, vat_known = amounts
        records.append(
            BillingRecord(
                date=when,
                client=(row.get(columns.get("client", ""), "") or "").strip(),
                amount_net=net,
                amount_gross=gross,
                vat_is_known=vat_known,
                document=(row.get(columns.get("document", ""), "") or "").strip(),
            )
        )
    return records, problems


def load_cashflow(path: str) -> tuple:
    """טוען קובץ תזרים. כיוון התנועה נקבע מעמודת סוג, ואם אין - מסימן הסכום."""
    headers, rows = read_rows(path)
    columns = build_column_map(headers)
    records, problems = [], []

    if "date" not in columns:
        raise ValueError(
            f"חסרה עמודת תאריך בקובץ התזרים. הכותרות שנמצאו: {', '.join(headers)}"
        )

    for index, row in enumerate(rows, start=2):
        when = parse_date(row.get(columns["date"], ""))
        amounts = split_vat(row, columns)
        if when is None or amounts is None:
            if _is_blank(row):
                continue
            problems.append(
                RowProblem("cashflow", index, "תאריך או סכום לא קריאים", _preview(row))
            )
            continue

        net, gross, vat_known = amounts
        direction = parse_direction(row.get(columns.get("direction", ""), ""))
        if direction is None:
            direction = EXPENSE if net < 0 else INCOME
        records.append(
            CashflowRecord(
                date=when,
                direction=direction,
                amount_net=abs(net),
                amount_gross=abs(gross),
                vat_is_known=vat_known,
                category=(row.get(columns.get("category", ""), "") or "").strip(),
                client=(row.get(columns.get("client", ""), "") or "").strip(),
            )
        )
    return records, problems


def _is_blank(row: dict) -> bool:
    return not any(str(value or "").strip() for value in row.values())


def _preview(row: dict) -> str:
    return " | ".join(f"{key}={value}" for key, value in list(row.items())[:4])
